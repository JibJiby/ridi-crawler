import re

import requests
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.load_workbook("data/data.xlsx")
ws = wb.active

ws['I1'] = "5점 비율"
ws['J1'] = "4점 비율"
ws['K1'] = "3점 비율"
ws['L1'] = "2점 비율"
ws['M1'] = "1점 비율"

for row in ws.iter_rows(min_row=2):
    url = row[7].value

    res = requests.get(url)
    soup = BeautifulSoup(res.text, 'html.parser')

    score_graph_list = soup.select("ul.score_graph li span.score_bar")
    # soup.select_one().get
    for i, el in enumerate(score_graph_list):
        width_value = re.search('[0-9]+', el.get('style'))
        width_value = width_value.group() if width_value is not None else 0
        print(str(5-i)+'점', width_value)
        ws.cell(row=int(row[0].value)+1, column=9+i, value=width_value)
        # row[8+i] = width_value + "%"
    print(row[0].value, "-"*50)


wb.save('./result/data_test_exp_rate.xlsx')