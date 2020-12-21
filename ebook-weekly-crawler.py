"""
베스트셀러만 분석.
판타지
e북 / 웹소설
각 베스트 셀러에는 1. 주간 베스트 2. 월간 베스트 3. 스테디셀러
필터는  대여, 성인, 비성인

최종적으로 분석하고자 하는 것은
    사람들은 어떤 것을 좋아하느냐
리디의 핵심 가치인 '고객 중심'으로 생각하자

1. 베스트 셀러의 공통점을 찾자
2. 어떤 태그가(?)가 많이 보일까
등등
"""

import re

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

# 쿼리로 페이지 바꾸는 형식
# 총 페이지 7페이지

def styling_worksheet(ws):
    FG_COLOR = 'ced4da' # 약한 회색
    # 순위 | 제목 | 장르 | 평점 | 평점준 수 | 권수 | 할인률 | 주소
    ws.column_dimensions['A'].width = 12 # 번호
    ws.column_dimensions['B'].width = 50 # 제목
    ws.column_dimensions['C'].width = 15 # 장르
    ws.column_dimensions['D'].width = 10 # 평점
    ws.column_dimensions['E'].width = 10 # 평점준 참가자 수
    ws.column_dimensions['F'].width = 10 # 권수
    ws.column_dimensions['G'].width = 10 # 할인률
    ws.column_dimensions['H'].width = 40 # 주소

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         bottom=Side(style='thin'),
                         top=Side(style='thin'))

    # 헤더 스타일
    for cell in ws.iter_cols(max_row=1, max_col=ws.max_column):
        cell[0].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=FG_COLOR)
        cell[0].border = thin_border
        cell[0].font = Font(bold=True)

    # 가운데 정렬
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

total_genre_dict = {}
DOMAIN_URL = "https://ridibooks.com"

wb = openpyxl.Workbook()
ws = wb.active

ws.append(['순위', '제목', '장르', '평점', '평가 수', '권 수', '할인률', '주소'])


for i in range(7):
    data = requests.get("https://ridibooks.com/bestsellers/fantasy"+f"?page={i+1}")
    soup = BeautifulSoup(data.text, 'html.parser')

    books_list = soup.select(".book_macro_110")


    for j, book in enumerate(books_list):
        if i==0 and j==0:
            # 맨 처음 상단 (광고?) 제외
            continue

        url = book.select_one(".title_link").get('href')
        title = book.select_one(".meta_title").text.strip()
        genre = book.select_one(".genre").text.strip()
        discount = ""
        if len(book.select(".discount_num")) == 1:
            # 할인이라면
            discount = book.select_one(".discount_num .num").text

        avg_point = "0"
        if len(book.select(".StarRate_Score")) != 0:
            avg_point = book.select_one(".StarRate_Score").text
        avg_point = avg_point.replace("점", "")

        num_rater = "0"
        if len(book.select(".StarRate_ParticipantCount")) != 0:
            num_rater = book.select_one(".StarRate_ParticipantCount").text
        num_rater = re.sub('[명,]', '',num_rater)

        num_series = "0"
        if len(book.select(".count_num")) != 0:
            num_series = book.select_one(".count_num").text
        # num_series = re.match('[0-9]+', num_series).group()
        if re.search('[0-9]+', num_series) is not None:
            num_series = re.search('[0-9]+', num_series).group()

        if i==0:
            number = j
        else:
            number = j+1

        # 순위 | 제목 | 장르 | 평점 | 평점준 수 | 권수 | 할인률 | 주소
        ws.append([30*i+number, title, genre, avg_point, num_rater, num_series, discount, DOMAIN_URL+url])
        print(30*i+number, genre, "  ||  ",title, "  ||  ", avg_point, "  ||  ", num_rater,"  ||  ",discount
              ,"  ||  ", num_series,"권", "  ||  ", DOMAIN_URL+url)
        total_genre_dict.setdefault(genre, 0)
        total_genre_dict[genre] += 1

print("장르 종류 : ",total_genre_dict)

styling_worksheet(ws) # 입력 다 하고, 정리
wb.save('test.xlsx')



